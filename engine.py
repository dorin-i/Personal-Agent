import io
from datetime import datetime
from openpyxl import load_workbook
import requests

GRAPH_BASE = "https://graph.microsoft.com/v1.0"

def download_excel(access_token, filename):
    url = f"{GRAPH_BASE}/me/drive/special/approot:/{filename}:/content"
    headers = {"Authorization": f"Bearer {access_token}"}
    r = requests.get(url, headers=headers)
    r.raise_for_status()
    return io.BytesIO(r.content)

def upload_excel(access_token, filename, file_bytes):
    url = f"{GRAPH_BASE}/me/drive/special/approot:/{filename}:/content"
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    }
    r = requests.put(url, headers=headers, data=file_bytes.getvalue())
    r.raise_for_status()

def insert_new_row_and_mark(access_token, filename):
    file_stream = download_excel(access_token, filename)
    wb = load_workbook(file_stream)
    ws = wb["INPUT_DAILY"]

    first_data_row = None
    for row in range(2, ws.max_row + 1):
        if ws[f"A{row}"].value:
            first_data_row = row
            break

    if not first_data_row:
        first_data_row = 2

    ws.insert_rows(first_data_row)

    today = datetime.now().strftime("%d.%m.%Y")
    ws[f"A{first_data_row}"] = today
    ws[f"B{first_data_row}"] = "TEST_INSERT"

    out_stream = io.BytesIO()
    wb.save(out_stream)
    out_stream.seek(0)

    upload_excel(access_token, filename, out_stream)

    return {
        "inserted_row": first_data_row,
        "date_written": today
    }
